import csv
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
import jdatetime

from .forms import TransactionForm
from .models import Budget, Category, RecurringTransaction, SavingsGoal, Transaction
from .services import generate_recurring_transactions

FA_TO_EN = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')


def _jmonth_range(jyear, jmonth):
    first = jdatetime.date(jyear, jmonth, 1).togregorian()
    if jmonth == 12:
        next_first = jdatetime.date(jyear + 1, 1, 1).togregorian()
    else:
        next_first = jdatetime.date(jyear, jmonth + 1, 1).togregorian()
    return first, next_first


def _persian_month_name(month):
    names = ['', 'فروردین', 'اردیبهشت', 'خرداد', 'تیر', 'مرداد', 'شهریور',
             'مهر', 'آبان', 'آذر', 'دی', 'بهمن', 'اسفند']
    return names[month]


def _parse_jdate(text):
    if not text:
        return None
    t = str(text).strip().translate(FA_TO_EN).replace('-', '/')
    parts = t.split('/')
    if len(parts) != 3:
        return None
    try:
        return jdatetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, TypeError):
        return None


def _last_jday(jyear, jmonth):
    first_g, next_g = _jmonth_range(jyear, jmonth)
    return (next_g - first_g).days


def _resolve_range(request):
    today_j = jdatetime.date.today()
    from_text = request.GET.get('from', '').strip()
    to_text = request.GET.get('to', '').strip()
    start_j = _parse_jdate(from_text)
    end_j = _parse_jdate(to_text)
    if start_j and end_j and start_j <= end_j:
        start = start_j.togregorian()
        end_excl = (end_j + timedelta(days=1)).togregorian()
        from_display = start_j.strftime('%Y/%m/%d')
        to_display = end_j.strftime('%Y/%m/%d')
    else:
        start_j = jdatetime.date(today_j.year, today_j.month, 1)
        start = start_j.togregorian()
        _, end_excl = _jmonth_range(today_j.year, today_j.month)
        end_j = jdatetime.date(today_j.year, today_j.month, _last_jday(today_j.year, today_j.month))
        from_display = start_j.strftime('%Y/%m/%d')
        to_display = end_j.strftime('%Y/%m/%d')
    return start, end_excl, from_display, to_display, start_j, end_j


@login_required
def dashboard(request):
    today_j = jdatetime.date.today()
    start, end_excl, from_display, to_display, start_j, end_j = _resolve_range(request)
    range_qs = request.user.transactions.filter(date__gte=start, date__lt=end_excl)
    income = range_qs.filter(type='income').aggregate(total=Sum('amount'))['total'] or 0
    expense = range_qs.filter(type='expense').aggregate(total=Sum('amount'))['total'] or 0
    recent = request.user.transactions.all()[:10]

    expense_categories = (range_qs.filter(type='expense')
                  .values('category__name', 'category__icon')
                  .annotate(total=Sum('amount'))
                  .order_by('-total')[:6])
    income_categories = (range_qs.filter(type='income')
                  .values('category__name', 'category__icon')
                  .annotate(total=Sum('amount'))
                  .order_by('-total')[:6])

    expense_pie = [{'name': c['category__name'], 'total': c['total']} for c in expense_categories]
    income_pie = [{'name': c['category__name'], 'total': c['total']} for c in income_categories]

    month_stats = []
    cursor = jdatetime.date.fromgregorian(date=start)
    last_j = jdatetime.date.fromgregorian(date=(end_excl - timedelta(days=1)))
    while cursor <= last_j:
        mf, ml = _jmonth_range(cursor.year, cursor.month)
        mf = max(mf, start)
        ml = min(ml, end_excl)
        mq = request.user.transactions.filter(date__gte=mf, date__lt=ml)
        inc = mq.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
        exp = mq.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0
        month_stats.append({'label': _persian_month_name(cursor.month), 'year': cursor.year,
                            'income': inc, 'expense': exp})
        if cursor.month == 12:
            cursor = jdatetime.date(cursor.year + 1, 1, 1)
        else:
            cursor = jdatetime.date(cursor.year, cursor.month + 1, 1)

    total_days = (end_excl - start).days
    step = 1 if total_days <= 31 else 7
    bins = {}
    for offset in range(0, total_days, step):
        bins[offset] = {'expense': 0, 'income': 0}
    for t in range_qs:
        offset = (t.date - start).days
        bin_offset = (offset // step) * step
        bins[bin_offset][t.type] += t.amount
    daily_data = []
    for offset in sorted(bins):
        d = start + timedelta(days=offset)
        j = jdatetime.date.fromgregorian(date=d)
        if step == 1:
            label = j.strftime('%m/%d')
        else:
            label = j.strftime('%m/%d') + ' ←'
        daily_data.append({'label': label, 'expense': bins[offset]['expense'], 'income': bins[offset]['income']})

    period_text = f'از {from_display} تا {to_display}'
    b_year, b_month = _parse_budget_month(request.GET, today_j)
    budget_alerts = [r for r in _budget_rows(request.user, b_year, b_month) if r['status'] in ('warn', 'over')]
    goal_rows = _goal_rows(request.user)
    context = {
        'income': income,
        'expense': expense,
        'balance': income - expense,
        'recent': recent,
        'budget_alerts': budget_alerts,
        'has_budgets': request.user.budgets.exists(),
        'b_year': b_year,
        'b_month': b_month,
        'budget_month_name': _persian_month_name(b_month),
        'budget_years': range(today_j.year - 1, today_j.year + 2),
        'months': range(1, 13),
        'goal_rows': goal_rows,
        'expense_categories': expense_categories,
        'income_categories': income_categories,
        'expense_pie': expense_pie,
        'income_pie': income_pie,
        'month_stats': month_stats,
        'daily_data': daily_data,
        'year': today_j.year,
        'month': today_j.month,
        'persian_month_name': _persian_month_name(today_j.month),
        'from_display': from_display,
        'to_display': to_display,
        'period_text': period_text,
    }
    return render(request, 'budget/dashboard.html', context)


@login_required
def transaction_list(request):
    generate_recurring_transactions(request.user)
    qs = request.user.transactions.select_related('category')
    year = request.GET.get('year') or ''
    month = request.GET.get('month') or ''
    cat = request.GET.get('category') or ''
    ttype = request.GET.get('type') or ''
    q = request.GET.get('q', '').strip()
    amount_min = request.GET.get('amount_min', '').strip()
    amount_max = request.GET.get('amount_max', '').strip()

    if year.isdigit() and month.isdigit():
        first, last = _jmonth_range(int(year), int(month))
        qs = qs.filter(date__gte=first, date__lt=last)
    elif year.isdigit():
        first = jdatetime.date(int(year), 1, 1).togregorian()
        last = jdatetime.date(int(year) + 1, 1, 1).togregorian()
        qs = qs.filter(date__gte=first, date__lt=last)
    if cat.isdigit():
        qs = qs.filter(category_id=int(cat))
    if ttype in ('income', 'expense'):
        qs = qs.filter(type=ttype)
    if q:
        qs = qs.filter(Q(note__icontains=q) | Q(category__name__icontains=q))
    if amount_min.isdigit():
        qs = qs.filter(amount__gte=int(amount_min))
    if amount_max.isdigit():
        qs = qs.filter(amount__lte=int(amount_max))

    total_income = qs.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = qs.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0
    today_j = jdatetime.date.today()
    context = {
        'transactions': qs,
        'categories': Category.objects.all(),
        'total_income': total_income,
        'total_expense': total_expense,
        'selected': {'year': year, 'month': month, 'category': cat, 'type': ttype,
                     'q': q, 'amount_min': amount_min, 'amount_max': amount_max},
        'years': range(today_j.year - 5, today_j.year + 1),
        'months': range(1, 13),
    }
    return render(request, 'budget/transaction_list.html', context)


@login_required
def export_transactions(request):
    fmt = request.GET.get('fmt', 'xlsx')
    qs = request.user.transactions.select_related('category').order_by('-date', '-created_at')

    from_text = request.GET.get('from', '').strip()
    to_text = request.GET.get('to', '').strip()
    start_j = _parse_jdate(from_text)
    end_j = _parse_jdate(to_text)
    if start_j and end_j and start_j <= end_j:
        qs = qs.filter(date__gte=start_j.togregorian(),
                       date__lt=(end_j + timedelta(days=1)).togregorian())
    else:
        year = request.GET.get('year') or ''
        month = request.GET.get('month') or ''
        if year.isdigit() and month.isdigit():
            first, last = _jmonth_range(int(year), int(month))
            qs = qs.filter(date__gte=first, date__lt=last)
        elif year.isdigit():
            first = jdatetime.date(int(year), 1, 1).togregorian()
            last = jdatetime.date(int(year) + 1, 1, 1).togregorian()
            qs = qs.filter(date__gte=first, date__lt=last)
    cat = request.GET.get('category') or ''
    ttype = request.GET.get('type') or ''
    if cat.isdigit():
        qs = qs.filter(category_id=int(cat))
    if ttype in ('income', 'expense'):
        qs = qs.filter(type=ttype)
    q = request.GET.get('q', '').strip()
    amount_min = request.GET.get('amount_min', '').strip()
    amount_max = request.GET.get('amount_max', '').strip()
    if q:
        qs = qs.filter(Q(note__icontains=q) | Q(category__name__icontains=q))
    if amount_min.isdigit():
        qs = qs.filter(amount__gte=int(amount_min))
    if amount_max.isdigit():
        qs = qs.filter(amount__lte=int(amount_max))

    total_income = qs.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = qs.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0

    headers = ['تاریخ', 'نوع', 'دسته', 'مبلغ (تومان)', 'توضیحات']
    rows = []
    for t in qs:
        rows.append([
            jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d'),
            t.get_type_display(),
            f'{t.category.icon} {t.category.name}',
            t.amount,
            t.note,
        ])
    rows.append(['', '', 'جمع درآمد', total_income, ''])
    rows.append(['', '', 'جمع هزینه', total_expense, ''])

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        return response

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'تراکنش‌ها'
    header_fill = PatternFill('solid', fgColor='0F766E')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
    bold_font = Font(bold=True)
    for r_i, row in enumerate(rows, start=2):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            if r_i >= len(rows):
                cell.font = bold_font
    for i, w in enumerate([12, 10, 24, 18, 32], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    return response


def _category_context():
    return {
        'income_categories': Category.objects.filter(type='income'),
        'expense_categories': Category.objects.filter(type='expense'),
    }


def _budget_rows(user, jyear=None, jmonth=None):
    if jyear is None or jmonth is None:
        today_j = jdatetime.date.today()
        jyear, jmonth = today_j.year, today_j.month
    first, last = _jmonth_range(jyear, jmonth)
    budgets = user.budgets.select_related('category')
    cat_ids = [b.category_id for b in budgets]
    spent_map = {r['category_id']: r['s'] for r in (
        user.transactions.filter(category_id__in=cat_ids, type='expense',
                                 date__gte=first, date__lt=last)
        .values('category_id').annotate(s=Sum('amount')))}
    rows = []
    for b in budgets:
        spent = spent_map.get(b.category_id, 0)
        pct = spent / b.amount * 100 if b.amount else 0
        rows.append({
            'budget': b,
            'spent': spent,
            'remaining': max(b.amount - spent, 0),
            'over_by': max(spent - b.amount, 0),
            'pct': min(pct, 100),
            'status': 'over' if spent > b.amount else ('warn' if pct >= 80 else 'ok'),
        })
    return rows


def _parse_budget_month(params, today_j):
    year = params.get('b_year', '')
    month = params.get('b_month', '')
    if year.isdigit() and month.isdigit():
        try:
            jdatetime.date(int(year), int(month), 1)
            return int(year), int(month)
        except (ValueError, TypeError):
            pass
    return today_j.year, today_j.month


def _goal_rows(user):
    today = date.today()
    rows = []
    for g in user.goals.all():
        pct = g.saved_amount / g.target_amount * 100 if g.target_amount else 0
        days_left = None
        jy = jm = jd = None
        if g.target_date:
            days_left = (g.target_date - today).days
            j = jdatetime.date.fromgregorian(date=g.target_date)
            jy, jm, jd = j.year, j.month, j.day
        rows.append({
            'goal': g,
            'pct': min(pct, 100),
            'remaining': max(g.target_amount - g.saved_amount, 0),
            'days_left': days_left,
            'done': g.saved_amount >= g.target_amount,
            'jy': jy,
            'jm': jm,
            'jd': jd,
            'has_date': g.target_date is not None,
        })
    return rows


def _parse_goal_date(request):
    y = request.POST.get('target_year', '').strip()
    m = request.POST.get('target_month', '').strip()
    d = request.POST.get('target_day', '').strip()
    if not y and not m and not d:
        return None
    if y.isdigit() and m.isdigit() and d.isdigit():
        try:
            return jdatetime.date(int(y), int(m), int(d)).togregorian()
        except (ValueError, TypeError):
            return None
    return None


ICON_GROUPS = [
    ('غذا و نوشیدنی', ['🍔', '🍕', '🍜', '🍛', '🍣', '🥗', '🍞', '🥛', '☕', '🧃', '🍎', '🍉', '🍓', '🍰', '🍩', '🍿', '🍗', '🥩', '🍤', '🥦', '🍚', '🧀', '🥤', '🍺', '🍷', '🍯']),
    ('خانه و مسکن', ['🏠', '🏢', '🏡', '🏬', '🛋', '🛏', '🚪', '🪟', '🧹', '🧺', '💡', '🪑', '🔑', '🚿', '🧻', '🪞', '🕯', '🌡']),
    ('حمل‌ونقل', ['🚗', '🚌', '🚕', '🏍', '🚲', '🚆', '✈️', '⛽', '🅿️', '🚙', '🚚', '🛵', '🚄', '🚇', '🚕', '🚢', '🚁']),
    ('تفریح و سرگرمی', ['🎮', '🎬', '🎵', '🎤', '🎯', '🏋️', '⚽', '🏀', '🎳', '🎨', '📺', '🎧', '🎪', '🎢', '🎰', '🎲', '🎭', '📷', '🎹', '🚴', '🏊', '⛺']),
    ('پوشاک', ['👕', '👖', '👗', '👟', '🧥', '🎩', '👜', '🧣', '🧤', '👞', '🩳', '👔', '👘', '🥿', '👒', '🧢', '👓']),
    ('بهداشت و سلامت', ['🩺', '💊', '🦷', '🧴', '🪥', '🏥', '🧬', '💉', '🚑', '🧼', '🦻', '👓', '💪']),
    ('تحصیل و کار', ['📚', '🎓', '✏️', '📝', '🖊', '💻', '📱', '🖥', '⌨️', '🧮', '📐', '📒', '🗄', '📁', '📂', '🖨', '📌']),
    ('درآمد و مالی', ['💰', '💵', '💳', '🏦', '💸', '📈', '💹', '🪙', '💎', '🏧', '💴', '🤑']),
    ('خانواده و شخصی', ['👨‍👩‍👧', '👶', '🧸', '💍', '🎁', '🎂', '🎉', '🕊', '💐', '🖼', '📦', '📿']),
    ('حیوانات', ['🐶', '🐱', '🐰', '🐹', '🐦', '🐠', '🐢', '🐴', '🐱', '🐼', '🦜', '🐝']),
    ('متفرقه', ['📦', '🛒', '🧾', '🗑', '🔧', '🪴', '🌳', '⚡', '📞', '🔋', '🧰', '🧯', '🔨', '🪫', '🌧', '☀️']),
]

GOAL_ICON_GROUPS = [
    ('سفر و تفریح', ['✈️', '🗺', '🏖', '🌊', '🏔', '⛺', '🧳', '🎒', '🏝', '⛷', '🎡', '🎢', '🏰', '🌋', '🏜', '🗽', '🕌', '🏯', '🌉', '🚢']),
    ('خودرو و حمل‌ونقل', ['🚗', '🚕', '🏍', '🚲', '🛵', '🚌', '🚆', '🛫', '⛽', '🅿️', '🚁', '🚀']),
    ('خانه و زندگی', ['🏠', '🏡', '🏢', '🛋', '🛏', '🪟', '🖼', '🪴', '💡', '🔑', '🛠', '🧰', '🚿', '🧺', '🕯', '🪞']),
    ('خرید و کالاها', ['🛒', '🛍', '🎁', '💍', '⌚', '📱', '💻', '🖥', '🎧', '📷', '🎮', '🎹', '📺', '🧥', '👗', '👟', '👜', '🎩', '🖨', '🔌']),
    ('پول و سرمایه', ['💰', '💵', '💳', '🏦', '💸', '📈', '💹', '🪙', '💎', '🏧', '🤑', '📊', '🧾']),
    ('تحصیل و یادگیری', ['🎓', '📚', '📝', '✏️', '🖊', '📐', '🧮', '🔬', '🧪', '📒', '🎼', '🎨', '📜', '🖋']),
    ('سلامت و خانواده', ['🩺', '💊', '🦷', '🏥', '👶', '🧸', '👨‍👩‍👧', '💐', '🎂', '🎉', '🎊', '💪', '🧘', '🏋️']),
    ('اشتراک و قبض', ['📱', '💻', '🎬', '🎵', '📺', '🌐', '☕', '🍕', '🍔', '🍜', '⚡', '🔔', '📦', '🍿']),
    ('متفرقه', ['🎯', '🏆', '🥇', '⭐', '🌟', '🔥', '❤️', '💙', '💚', '💜', '🎀', '🪄', '🌈', '☀️', '🌙', '🧭', '🔐', '🗝', '📂', '🕊', '🍀', '🦋', '🍃']),
]


@login_required
def transaction_add(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            t = form.save(commit=False)
            t.user = request.user
            t.save()
            return redirect('dashboard')
    else:
        form = TransactionForm(initial={'type': 'expense'})
    context = {'form': form, 'title': 'افزودن تراکنش'}
    context.update(_category_context())
    return render(request, 'budget/transaction_form.html', context)


@login_required
def transaction_edit(request, pk):
    t = get_object_or_404(Transaction, pk=pk, user=request.user)
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=t)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = TransactionForm(instance=t)
    context = {'form': form, 'title': 'ویرایش تراکنش'}
    context.update(_category_context())
    return render(request, 'budget/transaction_form.html', context)


@require_POST
@login_required
def transaction_delete(request, pk):
    t = get_object_or_404(Transaction, pk=pk, user=request.user)
    t.delete()
    return redirect('dashboard')


@login_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'budget/category_list.html', {
        'categories': categories,
        'icon_groups': ICON_GROUPS,
    })


@login_required
def category_edit(request, pk):
    c = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', '📁').strip()
        if name:
            c.name = name
            c.icon = icon
            c.save()
    return redirect('category_list')


@require_POST
@login_required
def category_add(request):
    name = request.POST.get('name', '').strip()
    ttype = request.POST.get('type', 'expense')
    icon = request.POST.get('icon', '📁')
    if name and ttype in ('income', 'expense'):
        Category.objects.create(name=name, type=ttype, icon=icon)
    return redirect('category_list')


@require_POST
@login_required
def category_delete(request, pk):
    get_object_or_404(Category, pk=pk).delete()
    return redirect('category_list')


@login_required
def budget_list(request):
    today_j = jdatetime.date.today()
    b_year, b_month = _parse_budget_month(request.GET, today_j)
    rows = _budget_rows(request.user, b_year, b_month)
    budgeted_ids = [r['budget'].category_id for r in rows]
    available = Category.objects.filter(type='expense').exclude(id__in=budgeted_ids)
    return render(request, 'budget/budget_list.html', {
        'rows': rows,
        'available': available,
        'year': b_year,
        'month': b_month,
        'persian_month_name': _persian_month_name(b_month),
        'b_year': b_year,
        'b_month': b_month,
        'budget_years': range(today_j.year - 1, today_j.year + 2),
        'months': range(1, 13),
    })


@require_POST
@login_required
def budget_add(request):
    cat_id = request.POST.get('category', '')
    amount = request.POST.get('amount', '').strip().translate(FA_TO_EN)
    if cat_id.isdigit() and amount.isdigit() and int(amount) > 0:
        category = Category.objects.filter(pk=int(cat_id), type='expense').first()
        if category and not request.user.budgets.filter(category=category).exists():
            Budget.objects.create(user=request.user, category=category, amount=int(amount))
    return redirect('budget_list')


@require_POST
@login_required
def budget_edit(request, pk):
    b = get_object_or_404(Budget, pk=pk, user=request.user)
    amount = request.POST.get('amount', '').strip().translate(FA_TO_EN)
    if amount.isdigit() and int(amount) > 0:
        b.amount = int(amount)
        b.save()
    return redirect('budget_list')


@require_POST
@login_required
def budget_delete(request, pk):
    b = get_object_or_404(Budget, pk=pk, user=request.user)
    b.delete()
    return redirect('budget_list')


@login_required
def recurring_list(request):
    generate_recurring_transactions(request.user)
    recurrings = request.user.recurrings.select_related('category')
    context = _category_context()
    context.update({
        'recurrings': recurrings,
        'days': range(1, 32),
    })
    return render(request, 'budget/recurring_list.html', context)


@require_POST
@login_required
def recurring_add(request):
    ttype = request.POST.get('type', 'expense')
    cat_id = request.POST.get('category', '')
    amount = request.POST.get('amount', '').strip().translate(FA_TO_EN)
    note = request.POST.get('note', '').strip()
    start_text = request.POST.get('start_date', '').strip()
    day = request.POST.get('day_of_month', '1').strip()
    start_j = _parse_jdate(start_text)
    if (ttype in ('income', 'expense') and cat_id.isdigit()
            and amount.isdigit() and int(amount) > 0
            and start_j and day.isdigit() and 1 <= int(day) <= 31):
        category = Category.objects.filter(pk=int(cat_id)).first()
        if category:
            RecurringTransaction.objects.create(
                user=request.user, type=ttype, category=category,
                amount=int(amount), note=note,
                start_date=start_j.togregorian(), day_of_month=int(day),
            )
    return redirect('recurring_list')


@require_POST
@login_required
def recurring_edit(request, pk):
    r = get_object_or_404(RecurringTransaction, pk=pk, user=request.user)
    amount = request.POST.get('amount', '').strip().translate(FA_TO_EN)
    note = request.POST.get('note', '').strip()
    day = request.POST.get('day_of_month', '').strip()
    if amount.isdigit() and int(amount) > 0 and day.isdigit() and 1 <= int(day) <= 31:
        r.amount = int(amount)
        r.note = note
        r.day_of_month = int(day)
        r.save()
    return redirect('recurring_list')


@require_POST
@login_required
def recurring_toggle(request, pk):
    r = get_object_or_404(RecurringTransaction, pk=pk, user=request.user)
    r.active = not r.active
    r.save(update_fields=['active'])
    return redirect('recurring_list')


@require_POST
@login_required
def recurring_delete(request, pk):
    r = get_object_or_404(RecurringTransaction, pk=pk, user=request.user)
    r.delete()
    return redirect('recurring_list')


@login_required
def goal_list(request):
    today_j = jdatetime.date.today()
    return render(request, 'budget/goal_list.html', {
        'goal_rows': _goal_rows(request.user),
        'icon_groups': GOAL_ICON_GROUPS,
        'goal_years': range(today_j.year - 1, today_j.year + 6),
        'months': range(1, 13),
        'days': range(1, 32),
        'default_y': today_j.year,
        'default_m': today_j.month,
        'default_d': today_j.day,
    })


@require_POST
@login_required
def goal_add(request):
    title = request.POST.get('title', '').strip()
    icon = request.POST.get('icon', '').strip() or '🎯'
    target = request.POST.get('target_amount', '').strip().translate(FA_TO_EN)
    target_date = _parse_goal_date(request)
    if title and target.isdigit() and int(target) > 0:
        SavingsGoal.objects.create(
            user=request.user, title=title, icon=icon,
            target_amount=int(target),
            target_date=target_date,
        )
    return redirect('goal_list')


@require_POST
@login_required
def goal_deposit(request, pk):
    g = get_object_or_404(SavingsGoal, pk=pk, user=request.user)
    amount = request.POST.get('amount', '').strip().translate(FA_TO_EN)
    if amount.isdigit() and int(amount) > 0:
        g.saved_amount += int(amount)
        g.save(update_fields=['saved_amount'])
    return redirect('goal_list')


@require_POST
@login_required
def goal_edit(request, pk):
    g = get_object_or_404(SavingsGoal, pk=pk, user=request.user)
    title = request.POST.get('title', '').strip()
    icon = request.POST.get('icon', '').strip() or '🎯'
    target = request.POST.get('target_amount', '').strip().translate(FA_TO_EN)
    saved = request.POST.get('saved_amount', '').strip().translate(FA_TO_EN)
    target_date = _parse_goal_date(request)
    if title and target.isdigit() and int(target) > 0 and saved.isdigit():
        g.title = title
        g.icon = icon
        g.target_amount = int(target)
        g.saved_amount = int(saved)
        g.target_date = target_date
        g.save()
    return redirect('goal_list')


@require_POST
@login_required
def goal_delete(request, pk):
    g = get_object_or_404(SavingsGoal, pk=pk, user=request.user)
    g.delete()
    return redirect('goal_list')


def service_worker(request):
    sw = Path(settings.BASE_DIR / 'budget' / 'static' / 'sw.js').read_text(encoding='utf-8')
    response = HttpResponse(sw, content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    response['Cache-Control'] = 'no-cache'
    return response
