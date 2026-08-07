from datetime import date
from io import BytesIO
from pathlib import Path

import jdatetime
from django.conf import settings
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from .models import Budget, Category, RecurringTransaction, SavingsGoal, Transaction
from .services import generate_recurring_transactions


def gj(year, month, day):
    return jdatetime.date(year, month, day).togregorian()


class BudgetAppTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='testuser', password='pass12345')
        self.expense_cat = Category.objects.create(name='خوراک', type='expense', icon='🍔')
        self.income_cat = Category.objects.create(name='حقوق', type='income', icon='💼')
        self.client.login(username='testuser', password='pass12345')

    def test_create_transaction_with_persian_date(self):
        response = self.client.post(reverse('transaction_add'), {
            'type': 'expense',
            'category': self.expense_cat.pk,
            'amount': 500000,
            'date': '1405/05/16',
            'note': 'خرید',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Transaction.objects.count(), 1)
        t = Transaction.objects.first()
        self.assertEqual(t.user, self.user)
        self.assertEqual(t.amount, 500000)
        self.assertEqual(t.date, gj(1405, 5, 16))

    def test_dashboard_shows_totals(self):
        Transaction.objects.create(user=self.user, type='income', category=self.income_cat,
                                   amount=1000000, date=gj(1405, 5, 1))
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=300000, date=gj(1405, 5, 2))
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '۷۰۰٬۰۰۰')

    def test_transaction_list_filter_by_persian_month(self):
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=100000, date=gj(1405, 5, 1))
        Transaction.objects.create(user=self.user, type='income', category=self.income_cat,
                                   amount=200000, date=gj(1405, 4, 1))
        response = self.client.get(reverse('transaction_list'), {'year': '1405', 'month': '5'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['transactions']), 1)

    def test_edit_transaction(self):
        t = Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                       amount=100000, date=gj(1405, 5, 1))
        response = self.client.post(reverse('transaction_edit', args=[t.pk]), {
            'type': 'expense',
            'category': self.expense_cat.pk,
            'amount': 999999,
            'date': '۱۴۰۵/۰۵/۱۰',
            'note': 'ویرایش شد',
        })
        self.assertEqual(response.status_code, 302)
        t.refresh_from_db()
        self.assertEqual(t.amount, 999999)
        self.assertEqual(t.date, gj(1405, 5, 10))

    def test_delete_transaction(self):
        t = Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                       amount=100000, date=date.today())
        response = self.client.post(reverse('transaction_delete', args=[t.pk]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Transaction.objects.count(), 0)

    def test_cannot_delete_others_transaction(self):
        other = User.objects.create_user(username='other', password='pass12345')
        t = Transaction.objects.create(user=other, type='expense', category=self.expense_cat,
                                       amount=100000, date=date.today())
        response = self.client.post(reverse('transaction_delete', args=[t.pk]))
        self.assertEqual(response.status_code, 404)

    def test_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_service_worker(self):
        response = self.client.get(reverse('service_worker'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/javascript')
        self.assertEqual(response['Service-Worker-Allowed'], '/')
        self.assertIn('addEventListener', response.content.decode())

    def test_manifest_served(self):
        manifest = Path(settings.BASE_DIR / 'budget' / 'static' / 'manifest.webmanifest')
        self.assertTrue(manifest.exists())
        self.assertIn('icon-192.png', manifest.read_text(encoding='utf-8'))

    def test_login_page_has_pwa_tags(self):
        self.client.logout()
        response = self.client.get(reverse('login'))
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('manifest.webmanifest', content)
        self.assertIn('/sw.js', content)

    def test_invalid_persian_date_rejected(self):
        response = self.client.post(reverse('transaction_add'), {
            'type': 'expense',
            'category': self.expense_cat.pk,
            'amount': 500000,
            'date': '1405/13/45',
            'note': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Transaction.objects.count(), 0)
        self.assertContains(response, 'نامعتبر')

    def test_dashboard_custom_date_range(self):
        Transaction.objects.create(user=self.user, type='income', category=self.income_cat,
                                   amount=1000000, date=gj(1405, 5, 1))
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=300000, date=gj(1405, 6, 10))
        response = self.client.get(reverse('dashboard'), {'from': '1405/05/01', 'to': '1405/05/31'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['expense'], 0)
        self.assertEqual(response.context['income'], 1000000)

    def test_dashboard_invalid_range_falls_back_to_current_month(self):
        response = self.client.get(reverse('dashboard'), {'from': '1405/05/20', 'to': '1405/05/01'})
        self.assertEqual(response.status_code, 200)
        today_j = jdatetime.date.today()
        self.assertEqual(response.context['from_display'], jdatetime.date(today_j.year, today_j.month, 1).strftime('%Y/%m/%d'))

    def test_budget_add(self):
        response = self.client.post(reverse('budget_add'), {
            'category': self.expense_cat.pk,
            'amount': '۵۰۰۰۰۰۰',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Budget.objects.count(), 1)
        self.assertEqual(Budget.objects.first().amount, 5000000)

    def test_budget_duplicate_category_rejected(self):
        Budget.objects.create(user=self.user, category=self.expense_cat, amount=1000000)
        self.client.post(reverse('budget_add'), {
            'category': self.expense_cat.pk,
            'amount': '2000000',
        })
        self.assertEqual(Budget.objects.count(), 1)

    def test_budget_list_shows_progress(self):
        Budget.objects.create(user=self.user, category=self.expense_cat, amount=5000000)
        today_j = jdatetime.date.today()
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=2000000, date=gj(today_j.year, today_j.month, 5))
        response = self.client.get(reverse('budget_list'))
        self.assertEqual(response.status_code, 200)
        row = response.context['rows'][0]
        self.assertEqual(row['spent'], 2000000)
        self.assertEqual(row['remaining'], 3000000)
        self.assertEqual(row['status'], 'ok')

    def test_budget_status_over_and_warn(self):
        Budget.objects.create(user=self.user, category=self.expense_cat, amount=1000000)
        today_j = jdatetime.date.today()
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=1200000, date=gj(today_j.year, today_j.month, 5))
        response = self.client.get(reverse('budget_list'))
        self.assertEqual(response.context['rows'][0]['status'], 'over')

    def test_budget_edit(self):
        b = Budget.objects.create(user=self.user, category=self.expense_cat, amount=1000000)
        self.client.post(reverse('budget_edit', args=[b.pk]), {'amount': '۳۰۰۰۰۰۰'})
        b.refresh_from_db()
        self.assertEqual(b.amount, 3000000)

    def test_budget_delete(self):
        b = Budget.objects.create(user=self.user, category=self.expense_cat, amount=1000000)
        self.client.post(reverse('budget_delete', args=[b.pk]))
        self.assertEqual(Budget.objects.count(), 0)

    def test_cannot_edit_others_budget(self):
        other = User.objects.create_user(username='other2', password='pass12345')
        b = Budget.objects.create(user=other, category=self.expense_cat, amount=1000000)
        response = self.client.post(reverse('budget_edit', args=[b.pk]), {'amount': '9999999'})
        self.assertEqual(response.status_code, 404)

    def test_dashboard_budget_alert_shows_when_over(self):
        Budget.objects.create(user=self.user, category=self.expense_cat, amount=1000000)
        today_j = jdatetime.date.today()
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=900000, date=gj(today_j.year, today_j.month, 5))
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['budget_alerts']), 1)
        self.assertEqual(response.context['budget_alerts'][0]['status'], 'warn')

    def test_budget_list_month_selection(self):
        Budget.objects.create(user=self.user, category=self.expense_cat, amount=5000000)
        today_j = jdatetime.date.today()
        past_y, past_m = today_j.year, today_j.month - 1
        if past_m == 0:
            past_y -= 1
            past_m = 12
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=1000000, date=gj(past_y, past_m, 10))
        response = self.client.get(reverse('budget_list'), {'b_year': past_y, 'b_month': past_m})
        self.assertEqual(response.context['rows'][0]['spent'], 1000000)
        self.assertEqual(response.context['rows'][0]['status'], 'ok')
        response_default = self.client.get(reverse('budget_list'))
        self.assertEqual(response_default.context['rows'][0]['spent'], 0)

    def test_dashboard_budget_alert_for_selected_month(self):
        Budget.objects.create(user=self.user, category=self.expense_cat, amount=1000000)
        today_j = jdatetime.date.today()
        past_y, past_m = today_j.year, today_j.month - 1
        if past_m == 0:
            past_y -= 1
            past_m = 12
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=1500000, date=gj(past_y, past_m, 10))
        response = self.client.get(reverse('dashboard'), {'b_year': past_y, 'b_month': past_m})
        self.assertEqual(len(response.context['budget_alerts']), 1)
        self.assertEqual(response.context['budget_alerts'][0]['status'], 'over')
        response_current = self.client.get(reverse('dashboard'))
        self.assertEqual(response_current.context['budget_alerts'], [])

    def test_export_csv(self):
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=250000, date=gj(1405, 5, 3), note='ناهار')
        response = self.client.get(reverse('transaction_export'), {'fmt': 'csv', 'year': '1405', 'month': '5'})
        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response['Content-Type'])
        body = response.content.decode('utf-8-sig')
        self.assertIn('تاریخ', body)
        self.assertIn('ناهار', body)
        self.assertIn('250000', body)

    def test_export_xlsx(self):
        Transaction.objects.create(user=self.user, type='income', category=self.income_cat,
                                   amount=1000000, date=gj(1405, 5, 3))
        response = self.client.get(reverse('transaction_export'), {'fmt': 'xlsx', 'year': '1405', 'month': '5'})
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws['A1'].value, 'تاریخ')
        self.assertEqual(ws['D2'].value, 1000000)
        self.assertEqual(ws['D3'].value, 1000000)
        self.assertEqual(ws['D4'].value, 0)

    def test_export_respects_date_range(self):
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=500000, date=gj(1405, 5, 3))
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=700000, date=gj(1405, 6, 3))
        response = self.client.get(reverse('transaction_export'),
                                   {'fmt': 'csv', 'from': '1405/05/01', 'to': '1405/05/31'})
        body = response.content.decode('utf-8-sig')
        self.assertIn('500000', body)
        self.assertNotIn('700000', body)

    def test_export_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('transaction_export'))
        self.assertEqual(response.status_code, 302)

    def test_search_by_note(self):
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=100000, date=date.today(), note='خرید ناهار')
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=50000, date=date.today(), note='بنزین')
        response = self.client.get(reverse('transaction_list'), {'q': 'ناهار'})
        self.assertEqual(len(response.context['transactions']), 1)
        self.assertEqual(response.context['transactions'][0].note, 'خرید ناهار')

    def test_search_by_category_name(self):
        other_cat = Category.objects.create(name='قهوه', type='expense', icon='☕')
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=100000, date=date.today(), note='x')
        Transaction.objects.create(user=self.user, type='expense', category=other_cat,
                                   amount=50000, date=date.today(), note='y')
        response = self.client.get(reverse('transaction_list'), {'q': 'قهوه'})
        self.assertEqual(len(response.context['transactions']), 1)
        self.assertEqual(response.context['transactions'][0].category.name, 'قهوه')

    def test_amount_range_filter(self):
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=100000, date=date.today())
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=500000, date=date.today())
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=900000, date=date.today())
        response = self.client.get(reverse('transaction_list'),
                                   {'amount_min': '200000', 'amount_max': '800000'})
        self.assertEqual(len(response.context['transactions']), 1)
        self.assertEqual(response.context['transactions'][0].amount, 500000)

    def test_combined_filters(self):
        t1 = Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                        amount=300000, date=gj(1405, 5, 10), note='سوپرمارکت')
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=300000, date=gj(1405, 6, 10), note='سوپرمارکت')
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=900000, date=gj(1405, 5, 10), note='سوپرمارکت')
        response = self.client.get(reverse('transaction_list'),
                                   {'q': 'سوپرمارکت', 'year': '1405', 'month': '5',
                                    'amount_min': '100000', 'amount_max': '500000'})
        self.assertEqual(len(response.context['transactions']), 1)
        self.assertEqual(response.context['transactions'][0].pk, t1.pk)

    def test_export_respects_search(self):
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=300000, date=date.today(), note='پیتزا')
        Transaction.objects.create(user=self.user, type='expense', category=self.expense_cat,
                                   amount=50000, date=date.today(), note='کرایه تاکسی')
        response = self.client.get(reverse('transaction_export'), {'fmt': 'csv', 'q': 'پیتزا'})
        body = response.content.decode('utf-8-sig')
        self.assertIn('پیتزا', body)
        self.assertNotIn('تاکسی', body)

    def test_recurring_add(self):
        response = self.client.post(reverse('recurring_add'), {
            'type': 'expense',
            'category': self.expense_cat.pk,
            'amount': '۸۰۰۰۰۰۰',
            'day_of_month': '1',
            'start_date': '1405/05/01',
            'note': 'اجاره خانه',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(RecurringTransaction.objects.count(), 1)
        r = RecurringTransaction.objects.first()
        self.assertEqual(r.amount, 8000000)
        self.assertEqual(r.start_date, gj(1405, 5, 1))

    def test_recurring_generates_transactions(self):
        Rec = RecurringTransaction.objects.create
        Rec(user=self.user, type='expense', category=self.expense_cat,
            amount=8000000, note='اجاره', start_date=gj(1405, 5, 1), day_of_month=1)
        today_j = jdatetime.date.today()
        created = generate_recurring_transactions(self.user,
                                                  up_to=gj(today_j.year, today_j.month, today_j.day))
        self.assertGreaterEqual(created, 1)
        self.assertEqual(Transaction.objects.filter(note='اجاره').count(), created)
        first_t = Transaction.objects.filter(note='اجاره').order_by('date').first()
        self.assertEqual(first_t.date, gj(1405, 5, 1))

    def test_recurring_no_duplicates(self):
        Rec = RecurringTransaction.objects.create
        Rec(user=self.user, type='expense', category=self.expense_cat,
            amount=8000000, note='اجاره', start_date=gj(1405, 5, 1), day_of_month=1)
        today_j = jdatetime.date.today()
        up_to = gj(today_j.year, today_j.month, today_j.day)
        generate_recurring_transactions(self.user, up_to=up_to)
        count1 = Transaction.objects.filter(note='اجاره').count()
        generate_recurring_transactions(self.user, up_to=up_to)
        self.assertEqual(Transaction.objects.filter(note='اجاره').count(), count1)

    def test_recurring_inactive_generates_nothing(self):
        Rec = RecurringTransaction.objects.create
        r = Rec(user=self.user, type='expense', category=self.expense_cat,
                amount=8000000, note='اشتراک', start_date=gj(1405, 4, 1), day_of_month=1)
        r.active = False
        r.save()
        today_j = jdatetime.date.today()
        generate_recurring_transactions(self.user, up_to=gj(today_j.year, today_j.month, today_j.day))
        self.assertEqual(Transaction.objects.filter(note='اشتراک').count(), 0)

    def test_recurring_list_page(self):
        Rec = RecurringTransaction.objects.create
        Rec(user=self.user, type='expense', category=self.expense_cat,
            amount=8000000, note='اجاره', start_date=gj(1405, 5, 1), day_of_month=1)
        response = self.client.get(reverse('recurring_list'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['recurrings']), 1)

    def test_recurring_edit(self):
        r = RecurringTransaction.objects.create(
            user=self.user, type='expense', category=self.expense_cat,
            amount=8000000, note='اجاره', start_date=gj(1405, 5, 1), day_of_month=1)
        self.client.post(reverse('recurring_edit', args=[r.pk]),
                         {'amount': '۹۰۰۰۰۰۰', 'day_of_month': '5', 'note': 'اجاره جدید'})
        r.refresh_from_db()
        self.assertEqual(r.amount, 9000000)
        self.assertEqual(r.day_of_month, 5)

    def test_recurring_toggle(self):
        r = RecurringTransaction.objects.create(
            user=self.user, type='expense', category=self.expense_cat,
            amount=8000000, note='اجاره', start_date=gj(1405, 5, 1), day_of_month=1)
        self.client.post(reverse('recurring_toggle', args=[r.pk]))
        r.refresh_from_db()
        self.assertFalse(r.active)
        self.client.post(reverse('recurring_toggle', args=[r.pk]))
        r.refresh_from_db()
        self.assertTrue(r.active)

    def test_recurring_delete(self):
        r = RecurringTransaction.objects.create(
            user=self.user, type='expense', category=self.expense_cat,
            amount=8000000, note='اجاره', start_date=gj(1405, 5, 1), day_of_month=1)
        self.client.post(reverse('recurring_delete', args=[r.pk]))
        self.assertEqual(RecurringTransaction.objects.count(), 0)

    def test_cannot_edit_others_recurring(self):
        other = User.objects.create_user(username='other3', password='pass12345')
        r = RecurringTransaction.objects.create(
            user=other, type='expense', category=self.expense_cat,
            amount=8000000, note='اجاره', start_date=gj(1405, 5, 1), day_of_month=1)
        response = self.client.post(reverse('recurring_edit', args=[r.pk]),
                                    {'amount': '9999999', 'day_of_month': '1', 'note': 'x'})
        self.assertEqual(response.status_code, 404)

    def test_goal_add(self):
        response = self.client.post(reverse('goal_add'), {
            'title': 'سفر به مشهد',
            'icon': '✈️',
            'target_amount': '۲۰۰۰۰۰۰۰',
            'target_year': '1405',
            'target_month': '12',
            'target_day': '1',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(SavingsGoal.objects.count(), 1)
        g = SavingsGoal.objects.first()
        self.assertEqual(g.target_amount, 20000000)
        self.assertEqual(g.target_date, gj(1405, 12, 1))

    def test_goal_add_without_date(self):
        response = self.client.post(reverse('goal_add'), {
            'title': 'بدون تاریخ',
            'target_amount': '5000000',
            'target_year': '',
            'target_month': '',
            'target_day': '',
        })
        self.assertEqual(response.status_code, 302)
        g = SavingsGoal.objects.first()
        self.assertIsNone(g.target_date)

    def test_goal_deposit(self):
        g = SavingsGoal.objects.create(user=self.user, title='هدف', target_amount=10000000)
        self.client.post(reverse('goal_deposit', args=[g.pk]), {'amount': '۲۰۰۰۰۰۰'})
        g.refresh_from_db()
        self.assertEqual(g.saved_amount, 2000000)

    def test_goal_list_shows_progress(self):
        SavingsGoal.objects.create(user=self.user, title='هدف', target_amount=10000000,
                                   saved_amount=4000000)
        response = self.client.get(reverse('goal_list'))
        self.assertEqual(response.status_code, 200)
        row = response.context['goal_rows'][0]
        self.assertEqual(row['pct'], 40)
        self.assertEqual(row['remaining'], 6000000)

    def test_goal_done_flag(self):
        SavingsGoal.objects.create(user=self.user, title='هدف', target_amount=1000000,
                                   saved_amount=1200000)
        response = self.client.get(reverse('goal_list'))
        self.assertTrue(response.context['goal_rows'][0]['done'])

    def test_goal_edit(self):
        g = SavingsGoal.objects.create(user=self.user, title='هدف', target_amount=1000000)
        self.client.post(reverse('goal_edit', args=[g.pk]), {
            'title': 'هدف جدید',
            'icon': '💰',
            'target_amount': '5000000',
            'saved_amount': '1000000',
            'target_year': '1405',
            'target_month': '11',
            'target_day': '30',
        })
        g.refresh_from_db()
        self.assertEqual(g.title, 'هدف جدید')
        self.assertEqual(g.target_amount, 5000000)
        self.assertEqual(g.saved_amount, 1000000)
        self.assertEqual(g.target_date, gj(1405, 11, 30))

    def test_goal_delete(self):
        g = SavingsGoal.objects.create(user=self.user, title='هدف', target_amount=1000000)
        self.client.post(reverse('goal_delete', args=[g.pk]))
        self.assertEqual(SavingsGoal.objects.count(), 0)

    def test_dashboard_shows_goals(self):
        SavingsGoal.objects.create(user=self.user, title='سفر', target_amount=1000000,
                                   saved_amount=500000)
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['goal_rows']), 1)
        self.assertEqual(response.context['goal_rows'][0]['pct'], 50)
